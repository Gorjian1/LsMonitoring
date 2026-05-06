"""Excel export — one sheet per node, with Status column."""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ..alarms import Status, evaluate
from ..config import AlarmConfig, Thresholds
from ..parser import Reading


HEADERS = ["Date-and-time", "Temperature", "A-axis", "B-axis", "A-variation", "B-variation", "Status", "Reasons"]

_FILLS = {
    Status.OK: None,
    Status.WARNING: PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid"),
    Status.CRITICAL: PatternFill(start_color="FFFF6B6B", end_color="FFFF6B6B", fill_type="solid"),
    Status.INVALID: PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid"),
}


def export_to_xlsx(
    out_path: Path,
    data: dict[int, Iterable[Reading]],
    thresholds: Thresholds,
    alarm: AlarmConfig,
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    for node_id, readings in data.items():
        ws = wb.create_sheet(title=f"Node-{node_id}"[:31])
        ws.append(HEADERS)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in readings:
            ev = evaluate(r, thresholds, alarm)
            row = [
                r.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                r.temperature,
                r.a_axis,
                r.b_axis,
                r.a_variation,
                r.b_variation,
                ev.status.value,
                "; ".join(ev.reasons) if ev.reasons else "",
            ]
            ws.append(row)
            fill = _FILLS.get(ev.status)
            if fill is not None:
                for cell in ws[ws.max_row]:
                    cell.fill = fill
        for col_idx, header in enumerate(HEADERS, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(14, len(header) + 2)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
